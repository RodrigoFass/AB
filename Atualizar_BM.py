################################
#                              #
#             BOT              #
#          ATUALIZAR           #
#                              #
################################

#region Import e misc

import win32api
import win32con
from PIL import ImageGrab
import time
import pytesseract
from pynput.keyboard import Controller
import openpyxl

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
path = 'C:\\Users\\rodfa\\OneDrive\\Área de Trabalho\\Tudo\\albion-tabela\\Tabela-BM.xlsx'
book = openpyxl.load_workbook(path)
sheet = book.sheetnames
keyboard = Controller()

#endregion

#region Funções

def clicar(x,y):

    time.sleep(0.3)
    win32api.SetCursorPos((x,y))
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x,y, 0, 0)
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x,y, 0, 0)
    return True

def extrair(nome,tier,lin):

    folh = book[sheet[pagina]]

    clicar(reset['x'],reset['y'])
    clicar(srch_bar['x'],srch_bar['y'])

    keyboard.type(nome)

    clicar(vender['x'],vender['y'])

    #region tiers

    if (tier == 1):

        col = 2

        for x in comandos:

            #region Determinantes

            if x[0]==5:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_1_5['x'],tier_1_5['y'])
                
            if x[0]==6:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_1_6['x'],tier_1_6['y'])
            
            if x[0]==7:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_1_7['x'],tier_1_7['y'])
                
            if x[0]==8:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_1_8['x'],tier_1_8['y'])
            
            #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

            if x[1]==0:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_0['x'],ench_0['y'])
                
            if x[1]==1:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_1['x'],ench_1['y'])
            
            if x[1]==2:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_2['x'],ench_2['y'])
                
            if x[1]==3:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_3['x'],ench_3['y'])

            #endregion

            time.sleep(0.3)

            ScrnSht = ImageGrab.grab(preco)
            valor = pytesseract.image_to_string(ScrnSht, config='--psm 7')


            valor = valor.replace('\n','')
            valor = valor.replace(',','')

            print(valor)

            folh.cell(row=lin,column=col).value = valor

            clicar(ench_bar['x'],ench_bar['y'])
            clicar(ench_0['x'],ench_0['y'])


            col +=1

    if (tier == 2):

        col = 2

        for x in comandos:

            #region Determinantes

            if x[0]==5:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_2_5['x'],tier_2_5['y'])
                
            if x[0]==6:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_2_6['x'],tier_2_6['y'])
            
            if x[0]==7:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_2_7['x'],tier_2_7['y'])
                
            if x[0]==8:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_2_8['x'],tier_2_8['y'])
            
            #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

            if x[1]==0:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_0['x'],ench_0['y'])
                
            if x[1]==1:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_1['x'],ench_1['y'])
            
            if x[1]==2:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_2['x'],ench_2['y'])
                
            if x[1]==3:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_3['x'],ench_3['y'])

            #endregion

            time.sleep(0.3)

            ScrnSht = ImageGrab.grab(preco)
            valor = pytesseract.image_to_string(ScrnSht, config='--psm 7')

            valor = valor.replace('\n','')
            valor = valor.replace(',','')

            print(valor)

            folh.cell(row=lin,column=col).value = valor

            clicar(ench_bar['x'],ench_bar['y'])
            clicar(ench_0['x'],ench_0['y'])

            col +=1

    if (tier == 3):

        col = 2

        for x in comandos:

            #region Determinantes

            if x[0]==5:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_3_5['x'],tier_3_5['y'])
                
            if x[0]==6:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_3_6['x'],tier_3_6['y'])
            
            if x[0]==7:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_3_7['x'],tier_3_7['y'])
                
            if x[0]==8:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_3_8['x'],tier_3_8['y'])
            
            #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

            if x[1]==0:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_0['x'],ench_0['y'])
                
            if x[1]==1:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_1['x'],ench_1['y'])
            
            if x[1]==2:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_2['x'],ench_2['y'])
                
            if x[1]==3:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_3['x'],ench_3['y'])

            #endregion

            time.sleep(0.3)

            ScrnSht = ImageGrab.grab(preco)
            valor = pytesseract.image_to_string(ScrnSht, config='--psm 7')


            valor = valor.replace('\n','')
            valor = valor.replace(',','')

            print(valor)

            folh.cell(row=lin,column=col).value = valor

            clicar(ench_bar['x'],ench_bar['y'])
            clicar(ench_0['x'],ench_0['y'])

            col +=1

    if (tier == 4):

        col = 2

        for x in comandos:

            #region Determinantes

            if x[0]==5:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_4_5['x'],tier_4_5['y'])
                
            if x[0]==6:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_4_6['x'],tier_4_6['y'])
            
            if x[0]==7:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_4_7['x'],tier_4_7['y'])
                
            if x[0]==8:

                clicar(tier_bar['x'],tier_bar['y'])
                clicar(tier_4_8['x'],tier_4_8['y'])
            
            #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

            if x[1]==0:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_0['x'],ench_0['y'])
                
            if x[1]==1:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_1['x'],ench_1['y'])
            
            if x[1]==2:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_2['x'],ench_2['y'])
                
            if x[1]==3:

                clicar(ench_bar['x'],ench_bar['y'])
                clicar(ench_3['x'],ench_3['y'])

            #endregion

            time.sleep(0.3)

            ScrnSht = ImageGrab.grab(preco)
            valor = pytesseract.image_to_string(ScrnSht, config='--psm 7')


            valor = valor.replace('\n','')
            valor = valor.replace(',','')
            
            print(valor)

            folh.cell(row=lin,column=col).value = valor

            clicar(ench_bar['x'],ench_bar['y'])
            clicar(ench_0['x'],ench_0['y'])

            col +=1

    #endregion

    book.save(path)

    clicar(fechar['x'],fechar['y'])
 
#endregion

#region Coordenadas

vender = {'x':1280,'y':435}
fechar = {'x':940,'y':315}
reset = {'x':720,'y':270}

ench_bar = {'x':665,'y':405}
tier_bar = {'x':520,'y':405}
srch_bar = {'x':640,'y':270}

preco = (1018, 368, 1094, 387)

#region Ench's

ench_0 = {'x':665,'y':430}
ench_1 = {'x':665,'y':460}
ench_2 = {'x':665,'y':490}
ench_3 = {'x':665,'y':510}

tier_1_5 = {'x':520,'y':540}
tier_1_6 = {'x':520,'y':565}
tier_1_7 = {'x':520,'y':590}
tier_1_8 = {'x':520,'y':620}

tier_2_5 = {'x':520,'y':510}
tier_2_6 = {'x':520,'y':540}
tier_2_7 = {'x':520,'y':565}
tier_2_8 = {'x':520,'y':590}

tier_3_5 = {'x':520,'y':480}
tier_3_6 = {'x':520,'y':510}
tier_3_7 = {'x':520,'y':540}
tier_3_8 = {'x':520,'y':565}

tier_4_5 = {'x':520,'y':454}
tier_4_6 = {'x':520,'y':480}
tier_4_7 = {'x':520,'y':510}
tier_4_8 = {'x':520,'y':540}

#endregion

comandos = [[5,2,10],[5,3,1],[6,0,30],[6,1,30],[6,2,15],[6,3,1],[7,0,30],[7,1,15],[7,2,1],[7,3,1],[8,0,5],[8,1,1],[8,2,1]]

#endregion

#region Main

print('''
Melee Ranged // Resto // Armaduras // Sair
     0            1          2         4
''')
pagina = int(input("insira oque deseja atualizar: "))

while(pagina!=4):

    time.sleep(2)

    linha = 2

    folha = book[sheet[pagina]]

    celula = (folha.cell(row=linha,column=29).value)
    tiers = int((folha.cell(row=linha,column=31).value))

    while (celula != None):

        extrair(celula,tiers,linha)

        linha += 1

        celula = (folha.cell(row=linha,column=29).value)
        tiers = int((folha.cell(row=linha,column=31).value))


    print('''
    Armaduras // Melee Ranged // Resto // Sair
        0             1            2        4
    ''')
    pagina = int(input("insira oque deseja atualizar: "))

#endregion
