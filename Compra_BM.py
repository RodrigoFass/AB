################################
#                              #
#             BOT              #
#           COMPRAR            #
#                              #
################################

#region Import e misc

from itertools import countEspada
import win32api
import win32con
from PIL import ImageGrab
import time
import pytesseract
from pynput.keyboard import Controller
import openpyxl
from threading import Thread, Event
import os

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
path = 'C:\\Users\\rodfa\\OneDrive\\Área de Trabalho\\Tudo\\albion-tabela\\Tabela-BM.xlsx'
book = openpyxl.load_workbook(path, data_only=True)
keyboard = Controller()

#endregion

#region Funções

def clicar(x,y):

    time.sleep(0.1)
    win32api.SetCursorPos((x,y))
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x,y, 0, 0)
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x,y, 0, 0)
    return True

def navegar(nome, folh, lin):


    abar = book[sheet[folh]]
    col = 15

    clicar(search_reset['x'],search_reset['y'])
    clicar(srchbar['x'],srchbar['y'])

    keyboard.type(nome)

    for x in comandos:

        if ((abar.cell(row=lin,column=30).value) != None) and (x[0]==8):

            clicar(search_reset['x'],search_reset['y'])
            clicar(srchbar['x'],srchbar['y'])
            keyboard.type(abar.cell(row=lin,column=30).value)
        
        col += 1

        #region Determinantes


        if x[0]==5:

            clicar(tierbar['x'],tierbar['y'])
            clicar(tier_5['x'],tier_5['y'])
            
        if x[0]==6:

            clicar(tierbar['x'],tierbar['y'])
            clicar(tier_6['x'],tier_6['y'])
        
        if x[0]==7:

            clicar(tierbar['x'],tierbar['y'])
            clicar(tier_7['x'],tier_7['y'])
            
        if x[0]==8:

            clicar(tierbar['x'],tierbar['y'])
            clicar(tier_8['x'],tier_8['y'])
        
        #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        if x[1]==0:

            clicar(enchbar['x'],enchbar['y'])
            clicar(ench_0['x'],ench_0['y'])
            
        if x[1]==1:

            clicar(enchbar['x'],enchbar['y'])
            clicar(ench_1['x'],ench_1['y'])
        
        if x[1]==2:

            clicar(enchbar['x'],enchbar['y'])
            clicar(ench_2['x'],ench_2['y'])
            
        if x[1]==3:

            clicar(enchbar['x'],enchbar['y'])
            clicar(ench_3['x'],ench_3['y'])
        
        #endregion


        auth = oferta()

        if (auth == True):            

            cellula = (abar.cell(row=lin,column=col)).value

            print(cellula)

            checar(cellula,x[2])

    return

def checar(velula,qntd):

    if velula == "#VALOR!":
        return

    for i in range (qntd):

        time.sleep(0.3)

        auth = oferta()

        if (auth == True):

            ScrnSht = ImageGrab.grab(preco)
            valor = pytesseract.image_to_string(ScrnSht, config='--psm 7')

            valor = valor.replace(',','')
            valor = valor.replace('\n','')

            try:
                valor = float(valor)
                print(valor,velula)

                if valor < float(velula):

                    clicar(buy_1['x'],buy_1['y'])
                    clicar(buy_2['x'],buy_2['y'])
            
                else:

                    return
            except:
                print('deu merda')
                return

    return

def oferta():

    time.sleep(0.5)

    ScrnSht = ImageGrab.grab(Sem_item)
    teste = pytesseract.image_to_string(ScrnSht, config='--psm 7')
    teste = teste.replace("\n", "")

    if teste == 'oferta':

        return False

    else:

        return True

#endregion

#region Coordenadas

tier_5 = {'x':960,'y':435}
tier_6 = {'x':960,'y':465}
tier_7 = {'x':960,'y':485}
tier_8 = {'x':960,'y':515}

ench_0 = {'x':1100,'y':325}
ench_1 = {'x':1100,'y':355}
ench_2 = {'x':1100,'y':375}
ench_3 = {'x':1100,'y':405}

buy_1 = {'x':1275,'y':430}
buy_2 = {'x':880,'y':735}

preco = (1070,412,1170,450)

tierbar = {'x':960,'y':270}
enchbar = {'x':1100,'y':270}
srchbar = {'x':640,'y':270}
search_reset = {'x':720,'y':270}
Sem_item = (970,450,1030,470)

comandos = [[5,2,6],[5,3,1],[6,0,20],[6,1,20],[6,2,8],[6,3,1],[7,0,20],[7,1,8],[7,2,1],[7,3,1],[8,0,3],[8,1,1],[8,2,1]]

#endregion

#region Itens

sheet = book.sheetnames

user_event = Event()
user_event.set()

def run_bot(sheet, user_event):
    for i in range (3):
        
        linha = 2
        aba = book[sheet[i]]

        celula = (aba.cell(row=linha,column=29).internal_value)

        while celula != None:

            navegar(celula, i, linha)

            linha += 1
            celula = (aba.cell(row=linha,column=29).internal_value)

        if user_event.is_set():
            pass
        else:
            os.system("pause")

#endregion

def pause_fun(sheet, user_event):
    while(True):
        p = input('p == pause, r == resume')
        p = str(p)
        if(p == 'p'):
            user_event.clear()
        if(p == 'r'):
            user_event.set()


bot_thread = Thread(target=run_bot, args=(sheet, user_event))
bot_thread.start()

user_thread = Thread(target=pause_fun, args=(sheet, user_event))
user_thread.start()
