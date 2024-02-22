import pyscreenshot as ImageGrab
import time
import pytesseract
import win32api
import win32con
from pynput.keyboard import Key, Controller
import openpyxl
import pydirectinput
import cv2
import numpy as np

book = openpyxl.load_workbook('tabela_BM.xlsx', data_only=True)
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'

price_region = (1098, 385, 1245, 430)
price_region_temp = (1098, 385, 1245, 430)
price_region_halberd = (1098, 415, 1240, 439)
reset_price = (662, 202)

tier = (966, 203)
tier_5 = (960, 408)
tier_6 = (960, 441)
tier_7 = (960, 477)
tier_8 = (960, 510)

enchant = (1151, 205)
enchant_0 = (1148, 270)
enchant_1 = (1148, 304)
enchant_2 = (1139, 339)
enchant_3 = (1141, 374)

search_bar = (564, 203)
keyboard = Controller()


def clicar(posicao):
    time.sleep(0.1)
    win32api.SetCursorPos((posicao[0], posicao[1]))
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, posicao[0], posicao[1], 0, 0)
    time.sleep(0.1)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, posicao[0], posicao[1], 0, 0)
    return True

def clicarCompra(posicao):
    time.sleep(0.2)
    win32api.SetCursorPos((posicao[0], posicao[1]))
    time.sleep(0.2)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, posicao[0], posicao[1], 0, 0)
    time.sleep(0.2)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, posicao[0], posicao[1], 0, 0)
    return True

def comprar_item():
    x = (1360, 415)
    clicarCompra(x) #clica em comprar
    y = (870, 781)
    clicarCompra(y) #compra

def clicarTipoDeArma(tipo):
    clicar((772, 201))
    if(tipo == "Melee"):
        clicar((775, 679))
    elif(tipo == "Ranged"):
        clicar((764, 848))
    elif(tipo == "Staffs"):
        clicar((767, 609))
    elif(tipo == "Armadura"):
        clicar((769, 304))


def extrair_num(ss_region):
    try:
        time.sleep(0.8)
        ss_img = ImageGrab.grab(bbox=ss_region)
        img = np.array(ss_img)
        img = cv2.resize(img, None, fx=1.2, fy=1.2, interpolation=cv2.INTER_CUBIC)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # th, img_gray_th_otsu = cv2.threshold(img_gray, 128, 192, cv2.THRESH_OTSU)
        valor2 = pytesseract.image_to_string(img, config='--psm 7 --oem 1 -c tessedit_char_whitelist=0123456789,')
        valor2 = valor2.strip()
        valor2 = valor2.split(".")
        valor2 = "".join(valor2)
        valor2 = valor2.split(",")
        valor2 = "".join(valor2)
        valor2 = valor2.lower()

        if (valor2.find("i") != -1):
            valor = valor2.replace("i", "1")
            return int(valor)
        elif (valor2.find("]") != -1):
            valor = valor2.replace("]", "1")
            return int(valor)
        elif (valor2.find("j") != -1):
            valor = valor2.replace("j", "1")
            return int(valor)
        elif (valor2.find("l") != -1):
            valor = valor2.replace("l", "1")
            return int(valor)
        elif (valor2.find(" ") != -1):
            valor = valor2.replace(" ", "")
            return int(valor)

        return int(valor2)
    except:
        return 100000

artefatos = ["Chama-corpos do", "Cajado de Monge Negro do", "Martelo Fúnebre do", "Mangual do", "Cajado Feiticeiro do", "Lâmina Aclarada do",
             "Armadura de Guarda-tumbas do", "Botas de Guarda-tumbas do", "Elmo de Guarda-tumbas do", "Casaco de Espreitador do",
             "Sapatos de Espreitador do", "Capuz de Espreitador do",
             "Capote de Druida do", "Robe de Druida do", "Sandálias de Druida do", "Elmo Real do", "Armadura Real do", "Botas Reais do", "Capuz Real do",
             "Casaco Real do", "Sapatos Reais do", "Capote Real do", "Robe Real do", "Sandálias Reais do", "Cajado Enregelante do",
             "Cajado Execrado do", "Cajado Incendiário do", "Cajado Druídico do", "Capa Demoníaca do", "Capa Morta-Viva do",
             "Capa de Morgana do", "Capa Protetora do", "Capa Herege do", "Sarcófago do", "Escudo do", "Brumário do", "Tocha do",
             "Olho dos Segredos do", "Tomo de Feitiços do", "Luvas Ursinas do", "Repetidor Lamentoso do",
             "Arco Sussurrante do", "Maça Pétrea do", "Cajado Avivador do", "Martelo Fúnebre do",
             "Luvas de Lutador do", "Braçadeiras de Batalha do", "Manoplas Cravadas do", "Elmo Real do",
             "Elmo de Guarda-tumbas do", "Armadura Real do", "Armadura de Guarda-tumbas do", "Botas Reais do",
             "Botas de Guarda-tumbas do", "Capuz Real do", "Capuz de Espreitador do", "Casaco Real do",
             "Casaco de Espreitador do", "Capote Real do", "Capote de Druida do", "Robe Real do", "Sandálias Reais do",
             "Sandálias de Druida do", "Maça do", "Arco do"]

qntd_item_normal = [0, 10, 0, 10, 30, 0, 15, 30, 35, 15]
qntd_item_artefato = [0, 0, 0, 0 ,0 ,0 ,0 ,0 ,0 ,0 ]
#[5, 0, 15, 15, 6, 1, 15, 5, 1, 0, 5, 1, 0]
qntd_item = []

tiers = [tier_8, tier_8, tier_7, tier_7, tier_7, tier_6, tier_6, tier_6, tier_6, tier_5]
enchants = [enchant_1, enchant_0, enchant_2, enchant_1, enchant_0, enchant_3, enchant_2, enchant_1, enchant_0, enchant_2]


tipo_item = int(
    input('Digite qual tipo de item é desejado:\n1- Armaduras\n2- Cajados\n3- Meeles\n4- Rangeds/OffHands\n'))
if tipo_item == 1:
    item = book['Armaduras']
    qntdLinhas = 26
    tipoArma = "Armadura"
    armaTroca = "vin"
elif tipo_item == 2:
    item = book['Cajados']
    qntdLinhas = 19
    tipoArma = "Staffs"
    tipoArma2 = "BooksEtc"
    armaTroca = "Tome Of Spells"
elif tipo_item == 3:
    item = book['Melee']
    qntdLinhas = 30
    tipoArma = "Melee"
    armaTroca = "Besta do"
elif tipo_item == 4:
    item = book["RangedOffHand"]
    tipoArma = "Ranged"
i = 2
total_lucro = 0
time.sleep(2)

clicarTipoDeArma(tipoArma)
while True:
    time.sleep(0.3)
    nome_i = item.cell(row=i, column=24).value
    if(nome_i == armaTroca):
        clicarTipoDeArma(tipoArma2)
    nome_item = (nome_i)
    print("Nome do item: " + str(nome_i))

    if nome_i in artefatos:
        qntd_item = qntd_item_artefato
    else:
        qntd_item = qntd_item_normal

    if nome_item == "halberd":
        price_region = price_region_halberd

    for x in range(1, 5):
        clicar(search_bar)
    keyboard.type(nome_item)
    j = 13
    tiersIndex = 0
    enchantsIndex = 0
    qntdItemIndex = 0
    print("Linha: " + str(i) )
    while True:
        clicar(tier)
        time.sleep(0.1)
        clicar(tiers[tiersIndex])
        clicar(enchant)
        time.sleep(0.1)
        clicar(enchants[enchantsIndex])
        time.sleep(0.1)
        pydirectinput.moveTo(enchant[0], enchant[1])
        time.sleep(0.1)
        price_bm = item.cell(row=i, column=j).value
        print("a :" + price_bm)
        if (price_bm != "#VALUE!"):
            price_bm = int(price_bm)
            qntdItensComprados = 0
            while qntdItensComprados < qntd_item[qntdItemIndex]:
                price_mkt = extrair_num(price_region) - 1000
                print("preço no BM: " + str(price_bm))
                print("preço no Market: " + str(price_mkt))
                if price_mkt <= price_bm:
                    comprar_item()
                    total_lucro += (int(item.cell(row=i, column=(j - 11) ).value) - price_mkt) - (price_mkt * 0.045)
                    print("Lucro até agora: " + str(total_lucro) + "\n")
                    qntdItensComprados += 1
                else:
                    qntdItensComprados = 40
        j += 1
        qntdItemIndex += 1
        tiersIndex += 1
        enchantsIndex += 1
        if (j == 23):
            break
    clicar(reset_price)
    i += 1
    price_region = price_region_temp
    if(i > qntdLinhas):
        print("Total de lucro: " + str(total_lucro))
        break